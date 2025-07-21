import openpyxl
import json
import os
import re


excel_path = 'شيت المنتجات.xlsx'
json_path = 'items.json'

table = openpyxl.load_workbook(excel_path, data_only=True)
sheet = table.active

header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

products = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    if not any(row): 
        continue
    product = dict(zip(header, row))
    product = {k.strip(): v for k, v in product.items() if k and v is not None}
    product = {k.strip(): v for k, v in product.items()}
    price_field = product.get('سعر المنتج', '')
    variants = []
    for size_label, size_key in [(r'large|Larg|كبير|كبيره|كبيرًا|كبيرا|كبيره', 'large'), (r'small|صغير|صغيره|صغيرًا|صغيرا|صغيره', 'small')]:
        match = re.search(rf'({size_label})\s*EGP\s*([\d\.,]+)', price_field, re.IGNORECASE)
        if match:
            price = match.group(2).replace('٫', '.').replace(',', '').strip()
            try:
                price = float(price)
            except Exception:
                pass
            variant = {**product, 'size': size_key, 'price': price}
            if 'تصنيف المنتج' in product:
                variant['تصنيف المنتج'] = product['تصنيف المنتج']
            variants.append(variant)
    if not variants:
        numbers = re.findall(r'([\d\.,]+)', price_field)
        if numbers:
            if len(numbers) == 2:
                try:
                    price_large = float(numbers[0].replace('٫', '.').replace(',', ''))
                    price_small = float(numbers[1].replace('٫', '.').replace(',', ''))
                except Exception:
                    price_large = numbers[0]
                    price_small = numbers[1]
                variant_large = {**product, 'size': 'large', 'price': price_large}
                variant_small = {**product, 'size': 'small', 'price': price_small}
                if 'تصنيف المنتج' in product:
                    variant_large['تصنيف المنتج'] = product['تصنيف المنتج']
                    variant_small['تصنيف المنتج'] = product['تصنيف المنتج']
                variants.append(variant_large)
                variants.append(variant_small)
            else:
                try:
                    price = float(numbers[0].replace('٫', '.').replace(',', ''))
                except Exception:
                    price = numbers[0]
                variant = {**product, 'size': 'unknown', 'price': price}
                if 'تصنيف المنتج' in product:
                    variant['تصنيف المنتج'] = product['تصنيف المنتج']
                variants.append(variant)
        else:
            variant = {**product, 'size': 'unknown'}
            if 'تصنيف المنتج' in product:
                variant['تصنيف المنتج'] = product['تصنيف المنتج']
            variants.append(variant)
    for v in variants:
        v.pop('سعر المنتج', None)
        v.pop('null', None)
    products.extend(variants)

with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(products, f, ensure_ascii=False, indent=2)

print(f"Cleaned and enhanced {len(products)} product variants to {json_path}") 